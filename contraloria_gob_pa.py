#!/usr/bin/env python
# coding: utf-8

import selenium # need to install
import time
import os
import datetime as dt
import pandas as pd # need to install
import shutil
from datetime import timedelta
from openpyxl import Workbook, load_workbook # need to install
from openpyxl.styles import Font
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import Select

def search(button): # clicks search
    print("Search Button clicked")
    button.click()
    print_time()

def add_institution(file_location): # add institution column to our downloaded files
    for file_name in os.listdir(file_location): # loop through files in the directory
        file_path = os.path.join(file_location, file_name) # create path to excel file
        wb = load_workbook(file_path) # open the file
        ws = wb.active # edit file
        ws['I5'] = "Institución" # add title
        ws['I5'].font = Font(bold = True, size = 12) # make the title bold and change size
        number_of_rows = ws.max_row # get number of rows filled
        institution_name = file_name[33:] # get the institution name
        
        for element in range(0, len(institution_name)): # parse institution names
            if institution_name[element] == '.':
                institution_name = institution_name[:element]
                break
                
        institution_name = institution_name.replace("_", " ")

        for row in range(6, number_of_rows+1): # assign institution names to the institution column
            position = "I" + str(row)
            ws[position] = institution_name
            ws[position].font = Font(size = 12)
        
        
        # realign headers
        ws.unmerge_cells("A1:H1") 
        ws.unmerge_cells("A2:H2")
        ws.unmerge_cells("E3:H3")
        ws.unmerge_cells("E4:H4")
        ws.unmerge_cells("A3:D3")

        for element in range(0,4): # delete first 4 rows
            ws.delete_rows(1)
            
        wb.save(file_path) # save file
        wb.close() # close file
        file_path = "" # reset file path

def combine_files(file_location): # combine our downloaded files
    add_institution(file_location) # add institution column to our sheets
    
    df_total = pd.DataFrame() 
    
    files = os.listdir(file_location) 
    
    for file in files: # loop through existing files in the directory
        if file.endswith('.xlsx'): # check if it's a sheet
            excel_file = pd.ExcelFile(f'{file_location}/{file}')
            sheets = excel_file.sheet_names
            
            for sheet in sheets:
                df = excel_file.parse(sheet_name = sheet)
                df_total = df_total.concat(df) # add the new data
                
            excel_file.close() # close the files
    
    for file in files: # delete unneeded files
        if file.endswith('.xlsx') and files != 'contraloria_gob_pa.xlsx' :
            os.remove(os.path.join(file_location, file)) 
            
    df_total.to_excel(f'{file_location}/contraloria_gob_pa.xlsx') # make new file with our data

def check_for_info(the_driver, text): # check if there is info stored in that option, a message box should appear if there isn't
    try:
        WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="modalNoRegistros"]/div/div/div[3]/button'))) # check for message box
        
    except: # timeout, meaning the message box did not appear
        print("Info available for " + text)
        has_info = True
        
    else: # message box appeared
        has_info = False
        print("No info available for " + text)
        accept = driver.find_element(By.XPATH, '//*[@id="modalNoRegistros"]/div/div/div[3]/button') # find accept
        accept.click() # click accept button
        
    return has_info

def select_drop_down_and_search(the_driver, the_text):
    drop = Select(the_driver.find_element(By.XPATH,"//*[@id=\"MainContent_ddlInstituciones\"]")) # select the drop down menu
    drop.select_by_visible_text(the_text) # choose an option in the drop down menu
    
    search_button = driver.find_element(By.ID, "MainContent_btnBuscar") # find the search button
    search(search_button) # clicks search

def get_options(driver):
    select = Select(driver.find_element(By.ID, "MainContent_ddlInstituciones")) # select the id where the options are located    
    options = select.options # select all of the available options under that ID and store them in the options list
    
    menu_text = [] # intialize text array
    
    for option in options: # create array of menu text
        menu_text.append(option.text) # add text to the array
        
    menu_text.pop(0) # remove the first one since it's the default value
    
    return menu_text

def download_files(driver, file_location): # download files from site
    file_downloaded = False
    
    # get number of files in folder before download
    pre_download = os.listdir(file_location) 
    pre_download_size = len(pre_download)
    
    WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="Btn_Descargar"]'))) # wait at most 1 minute for the download button to appear
    # some of the files are very large and take a while to download
    
    # download button was found
    download = driver.find_element(By.ID, "Btn_Descargar") # find the download button
    download.click() # click the download button
        
    post_download = os.listdir(file_location) # get number of files in folder after download
    post_download_size = len(post_download)
    if post_download_size == pre_download_size + 1: # check if file was downloaded
        file_downloaded = True
        
    return file_downloaded

def print_time():
    dt.datetime.now().time()
    print(dt.datetime.now().time())

def elapsed_time(start_time, end_time):
    total_time_in_seconds = end_time - start_time
    total_time = str(dt.timedelta(seconds = total_time_in_seconds))
    return total_time

# main
if __name__ == "__main__":
    # set up download folder
    download_directory = os.path.join("data", "contraloria_gob_pa") # change this to be the location of your downloaded files with backslashes
    shutil.rmtree(download_directory, ignore_errors=True)
    os.makedirs(download_directory)
    
    url = 'https://www.contraloria.gob.pa/CGR.PLANILLAGOB.UI/Formas/Index'
    
    print("Starting code")

    start_time = time.time() # start timer
    print_time() # print starting time
    
    # Set up driver
    chrome_options = webdriver.ChromeOptions() 

    # Change our download destination
    chrome_preferences = {
        "download.default_directory": os.path.abspath(download_directory)
    }

    chrome_options.add_experimental_option("prefs",chrome_preferences) # add it to preferences
    
    # Finish setting up driver
    chrome_options.headless = True
    driver = webdriver.Chrome(options = chrome_options)
    driver.set_page_load_timeout(2000) # change default page load from 300 seconds to 2000 seconds
    try:
        driver.get(url)
    except:
        print("No internet connection")
        driver.quit()
        quit()

    time.sleep(2)

    menu_text = get_options(driver) # create list of options from the drop down menu
    number_of_files = 0
    number_of_files_expected = len(menu_text)
    error_text = [] # list of options that didn't download the first time
    print(str(number_of_files_expected) + " Files Expected")

    # initial loop through the files
    for text in menu_text: # loop through each of the options/institutions
        file_downloaded = False # reset
        try:
            print(text) 
            print_time() 

            select_drop_down_and_search(driver,text) # select the option in the drop down menu and click search

            has_info = check_for_info(driver, text) # check if there is info stored in that option
            
            if has_info: # if there is available info
                while not file_downloaded: # while the file hasn't been downloaded
                    print("Attempting to download " + text)
                    file_downloaded = download_files(driver, download_directory) # download info

            else:
                number_of_files_expected-=1 # decrease the amount of files expected
                
        except: # time out exception
            time.sleep(3)
            driver.get(url) # reload page
            print("Error for " + text) # load error
            error_text.append(text)
        
        if file_downloaded:
            number_of_files += 1
            print(text + " Downloaded")
        
        time.sleep(5) # give the site some time to rest

        select = Select(driver.find_element(By.ID, "MainContent_ddlInstituciones")) # select the id where the options are located
        has_info = False # reset has_info

    # let's try downloading the files there were errors for

    error_attempts = 0 
    max_error_attempts = len(error_text) * 2 # max tries we will try to download the ones that didn't download the first time

    if number_of_files != number_of_files_expected:
        print("Beginning extra attempts to download files that did not download")
        
    while number_of_files != number_of_files_expected: # loop through each file that didn't download
        file_downloaded = False # reset
        try:
            print(error_text[0]) 
            print_time() 

            select_drop_down_and_search(driver,error_text[0]) # select the option in the drop down menu and click search

            has_info = check_for_info(driver, error_text[0]) # check if there is info stored in that option
            
            if has_info: # if there is available info
                while not file_downloaded: # while the file hasn't been downloaded
                    print("Attempting to download " + error_text[0])
                    file_downloaded = download_files(driver, download_directory) # download info

            else:
                number_of_files-=1 # decrease the amount of files expected
                
        except: # time out exception
            time.sleep(3)
            driver.get(url) # reload page
            print("Error for " + error_text[0]) # load error
        
        if file_downloaded:
            print(error_text[0] + " Downloaded")
            error_text.pop(0) # remove that error from list of errors
            number_of_files+=1 # increase number of files downloaded
            
        time.sleep(5) # give the site some time to rest

        select = Select(driver.find_element(By.ID, "MainContent_ddlInstituciones")) # select the id where the options are located
        has_info = False # reset has_info
        error_attempts += 1 # increase error attempts
        print(str(error_attempts) + " attempts to download error files")
        if error_attempts > max_error_attempts: # check if we have exceeded max attempts... time to give up
            print("Could not download " + error_text[0])
            break
    driver.quit()

    print(str(number_of_files) + " Files Downloaded")
    print(str(number_of_files_expected) + " Files Expected")
    combine_files(download_directory) # combine our files

    print_time()
    end_time = time.time() # stop timer
    total_time = elapsed_time(start_time, end_time) # calculate total time elapsed
    print("Total Time Elapsed: " + total_time)

    quit()
